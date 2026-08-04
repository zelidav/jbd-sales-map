#!/usr/bin/env python3
"""Brand-agnostic, product-led targeting layer.

Replaces the old two-brand model (JB=premium / Dragonfly=value) with per-category
price positioning, so a rep selling ANY product can find the doors that move that
kind of product at that kind of price.

For every store x category it attaches volume, units, avg price and a price tier
(Value / Mid / Prem), plus one overall customer-quality tier (H / M / L).

Usage:
  python tools/build_category_fit.py \
      --all   ~/Downloads/PULL_NY_all_1mo.xlsx ~/Downloads/PULL_NY_all_2mo.xlsx \
      --cat   flower=~/Downloads/PULL_cat_flower.xlsx \
      --cat   joint=~/Downloads/PULL_cat_joint.xlsx  ...

--all takes the short and mid ALL-CATEGORY windows (nested) and is used for the
quality tier's volume and momentum components.
"""
import os, re, sys, json, statistics, unicodedata
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
HTML = os.path.join(ROOT, "index.html")
CACHE = os.path.join(ROOT, "tools", "_cache")
STOP = set("the llc inc co of a and an at to ny nyc rec dispensary dispensaries store shop adult use".split())


def cset(s):
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().lower()
    s = re.sub(r"\(.*?\)", " ", s)
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return frozenset(t for t in s.split() if t and t not in STOP)


def cbase(s):
    return cset(str(s or "").split(" - ")[0])


def window_days(avgs):
    """True window length: 'Avg # SKUs Stocked' is a per-day mean, so its denominator
    is the day count. Exports carry no date metadata — never guess."""
    frac = [v for v in avgs if isinstance(v, (int, float)) and v != int(v)]
    if not frac:
        return None
    for d in range(1, 200):
        if sum(1 for v in frac if abs(v * d - round(v * d)) < 1e-3) / len(frac) > 0.85:
            return d
    return None


def load_rank(path, strict=True):
    wb = openpyxl.load_workbook(os.path.expanduser(path), data_only=True)
    ws = wb[wb.sheetnames[0]]
    out, avgs = {}, []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        out[r[1]] = {"rank": int(r[0]), "units": int(r[2] or 0),
                     "price": float(r[3] or 0), "vol": int(r[4] or 0)}
        if len(avgs) < 120 and len(r) > 5:
            avgs.append(r[5])
    days = window_days(avgs)
    if not days and strict:
        # The --all windows drive the momentum maths, so a wrong span there is corrupting.
        raise SystemExit(f"Cannot determine window length for {os.path.basename(path)} — refusing to guess.")
    return out, days


def pct_rank(sorted_vals, x):
    """Fraction of values at or below x, 0..1."""
    if not sorted_vals:
        return 0.0
    lo = 0
    for v in sorted_vals:
        if v <= x:
            lo += 1
        else:
            break
    return lo / len(sorted_vals)


def tier_from_pct(p):
    return "Prem" if p >= 0.67 else ("Mid" if p >= 0.33 else "Value")


def main():
    allf, cats = [], {}
    i = 1
    argv = sys.argv[1:]
    while i <= len(argv):
        a = argv[i - 1]
        if a == "--all":
            while i < len(argv) and not argv[i].startswith("--"):
                allf.append(argv[i]); i += 1
        elif a == "--cat":
            spec = argv[i]; i += 1
            name, _, path = spec.partition("=")
            cats[name] = path
        i += 1

    if len(allf) < 2 or not cats:
        raise SystemExit("Need --all <short> <mid> and at least one --cat name=file.")

    loaded = sorted((load_rank(f) + (f,) for f in allf), key=lambda t: t[1])
    (dS, nS, fS), (dM, nM, fM) = loaded[0], loaded[1]
    print(f"all-category windows: {nS}d {os.path.basename(fS)} | {nM}d {os.path.basename(fM)}")
    rates = [sum(x['vol'] for x in dS.values()) / nS, sum(x['vol'] for x in dM.values()) / nM]
    if max(rates) / min(rates) > 1.35:
        raise SystemExit("The two --all files disagree on $/day — different cuts, not different windows.")

    html = open(HTML, encoding="utf-8").read()
    m = re.search(r"var DATA=(\[.*?\]);", html, re.S)
    data = json.loads(m.group(1))

    ocm = {}
    p = os.path.join(CACHE, "ocm_full.json")
    if os.path.exists(p):
        ocm = json.load(open(p))

    index = {}
    for d in data:
        x = ocm.get(d.get("lic")) or {}
        for nm in (d.get("n"), x.get("dba"), x.get("ent")):
            for v in (cset(nm), cbase(nm)):
                if v:
                    index.setdefault(v, d)

    def door(store):
        for k in (cset(store), cbase(store)):
            if k in index:
                return index[k]
        return None

    # ---- customer quality tier: 50% volume pct + 30% price pct + 20% momentum vs market
    vols = sorted(x["vol"] for x in dM.values())
    prices = sorted(x["price"] for x in dM.values() if x["price"] > 0)
    moms = {}
    for st, rM in dM.items():
        rS = dS.get(st)
        if rS and rM["vol"] > rS["vol"]:
            prior = (rM["vol"] - rS["vol"]) / (nM - nS)
            if prior > 0:
                moms[st] = 100 * ((rS["vol"] / nS) / prior - 1)
    med_mom = statistics.median(moms.values()) if moms else 0
    mom_vals = sorted(moms.values())

    for d in data:
        d.pop("qt", None); d.pop("qs", None); d.pop("cat", None)

    tiered = 0
    for st, rM in dM.items():
        d = door(st)
        if not d:
            continue
        vp = pct_rank(vols, rM["vol"])
        pp = pct_rank(prices, rM["price"]) if rM["price"] > 0 else 0.5
        mp = pct_rank(mom_vals, moms[st]) if st in moms else 0.5
        score = round(100 * (0.50 * vp + 0.30 * pp + 0.20 * mp))
        d["qs"] = score
        d["qt"] = "H" if score >= 70 else ("M" if score >= 40 else "L")
        tiered += 1

    # ---- per-category volume / units / price / price-tier
    summary = {}
    for name, path in cats.items():
        rows, days = load_rank(path, strict=False)
        cprices = sorted(x["price"] for x in rows.values() if x["price"] > 0)
        n = 0
        for st, r in rows.items():
            d = door(st)
            if not d or r["vol"] <= 0:
                continue
            t = tier_from_pct(pct_rank(cprices, r["price"])) if r["price"] > 0 else "Mid"
            d.setdefault("cat", {})[name] = {
                "v": r["vol"], "u": r["units"], "p": round(r["price"], 2), "t": t,
            }
            n += 1
        tot = sum(x["vol"] for x in rows.values())
        summary[name] = {"days": days or "n/a", "stores": n, "total": tot,
                         "median_price": round(statistics.median(cprices), 2) if cprices else None}
        print(f"  {name:12} {str(days or '?'):>4}d stores={n:>4}  ${tot:>12,}  median price "
              f"${summary[name]['median_price']}")

    json.dump({"categories": summary, "quality": {"median_mom": round(med_mom, 1)}},
              open(os.path.join(CACHE, "category_fit.json"), "w"), indent=1)

    qt = {}
    for d in data:
        if d.get("qt"):
            qt[d["qt"]] = qt.get(d["qt"], 0) + 1
    print(f"\nquality tiers: {qt}  (scored {tiered} doors)")

    if "--dry" in sys.argv:
        print("--dry: no write."); return

    cj = json.dumps(sorted(cats.keys()))
    new = html[:m.start(1)] + json.dumps(data, ensure_ascii=False) + html[m.end(1):]
    if re.search(r"var CATS=\[.*?\];", new):
        new = re.sub(r"var CATS=\[.*?\];", f"var CATS={cj};", new, count=1)
    else:
        new = new.replace("var DATA=", f"var CATS={cj};\nvar DATA=", 1)
    open(HTML, "w", encoding="utf-8").write(new)
    print("Wrote per-category fit + quality tiers into index.html")


if __name__ == "__main__":
    main()
