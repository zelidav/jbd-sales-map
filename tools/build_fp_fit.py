#!/usr/bin/env python3
"""Flower+Preroll placement intelligence: data-driven JB premium tiers + Dragonfly value fit.

Two brands, opposite ends, flagged separately on every store:
  - Jerome Baker (JB) = PREMIUM (7g jars $65-75, 1g prerolls $13-14 retail).
    Belongs where premium flower/preroll sells. JB fit favors high avg F+PR price,
    real F+PR volume, healthy relative velocity, strong Pistil decile.
  - Dragonfly (Dfly) = VALUE (cheapest in every category). Belongs where value/volume
    flower/preroll moves. Dfly fit favors high F+PR UNIT volume at a value price.

Inputs: the Flower+Preroll store-rank exports (30/90-day; auto-detected by total).
Joins to map DATA by name. Attaches per store:
  fpvol (90d $), fpunits (90d), fpprice (avg $), fpvel (velocity vs market median),
  jbfit (0-100), jbt (1/2/3 or absent), dflyfit (0-100), dft (1/2/3 or absent).

Tiers (transparent cutoffs on the fit score, among qualifiers):
  T1 >= 70, T2 50-69, T3 35-49.  Below 35 = not a target for that brand.

Usage: python tools/build_fp_fit.py FP_A.xlsx FP_B.xlsx        # the two F+PR store files
       python tools/build_fp_fit.py                            # auto: 2 smallest-total store_rank in ~/Downloads
"""
import re, os, sys, json, glob, statistics, unicodedata
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
HTML = os.path.join(ROOT, "index.html")
STOP = set("the llc inc co of a and an at to ny nyc rec dispensary dispensaries store shop adult use".split())


def cset(s):
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().lower()
    s = re.sub(r"\(.*?\)", " ", s); s = re.sub(r"[^a-z0-9 ]", " ", s)
    return frozenset(t for t in s.split() if t and t not in STOP)


def cbase(s):
    return cset(str(s or "").split(" - ")[0])


def loadstore(p):
    wb = openpyxl.load_workbook(p, data_only=True); ws = wb[wb.sheetnames[0]]
    return {r[1]: {"rank": int(r[0]), "units": int(r[2] or 0), "avgp": float(r[3] or 0), "vol": int(r[4] or 0)}
            for r in ws.iter_rows(min_row=2, values_only=True) if r[0] is not None}


def pctile_fn(vals):
    s = sorted(vals)
    n = len(s)
    def f(x):
        lo = 0
        for v in s:
            if v <= x: lo += 1
            else: break
        return lo / n if n else 0
    return f


def tier(score):
    return 1 if score >= 70 else (2 if score >= 50 else (3 if score >= 35 else None))


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if len(args) >= 2:
        files = args[:2]
    else:
        # the two smallest-total store_rank files = the Flower+Preroll cut (30d + 90d)
        cand = sorted(glob.glob(os.path.join(os.path.expanduser("~"), "Downloads", "store_rank_*.xlsx")),
                      key=lambda p: sum(x["vol"] for x in loadstore(p).values()))
        files = cand[:2]
    a, b = loadstore(files[0]), loadstore(files[1])
    fp30, fp90 = (a, b) if sum(x["vol"] for x in a.values()) <= sum(x["vol"] for x in b.values()) else (b, a)
    print(f"F+PR 30-day: {os.path.basename(files[0] if a is fp30 else files[1])}  ${sum(x['vol'] for x in fp30.values()):,}")
    print(f"F+PR 90-day: {os.path.basename(files[1] if b is fp90 else files[0])}  ${sum(x['vol'] for x in fp90.values()):,}")

    vel = {s: round(100 * (fp30[s]["vol"] / (v["vol"] / 3.0) - 1)) for s, v in fp90.items() if v["vol"] > 0 and s in fp30}
    medv = statistics.median(vel.values())

    html = open(HTML, encoding="utf-8").read()
    m = re.search(r"var DATA=(\[.*?\]);", html, re.S)
    data = json.loads(m.group(1))

    # index F+PR stores by FULL normalized name (location-specific) so chains like
    # "Curaleaf - Queens" vs "Curaleaf - Hudson Valley" don't collapse. Base-name
    # fallback only when that base is unique in the F+PR file.
    fp_full = {}
    base_count = {}
    fp_base = {}
    for s, v in fp90.items():
        fp_full.setdefault(cset(s), (s, v))
        b = cbase(s)
        base_count[b] = base_count.get(b, 0) + 1
        fp_base.setdefault(b, (s, v))
    fp_base = {b: sv for b, sv in fp_base.items() if base_count[b] == 1}

    def fp_lookup(name):
        return fp_full.get(cset(name)) or fp_base.get(cbase(name))

    # percentile functions over the F+PR universe
    pP = pctile_fn([v["avgp"] for v in fp90.values() if v["avgp"] > 0])
    pV = pctile_fn([v["vol"] for v in fp90.values()])
    pU = pctile_fn([v["units"] for v in fp90.values()])
    relv = {s: vel.get(s, medv) - medv for s in fp90}
    velvals = list(relv.values())
    pVel = pctile_fn(velvals)

    used = set()
    matched = 0
    for d in data:
        for k in ("fpvol", "fpunits", "fpprice", "fpvel", "jbfit", "jbt", "dflyfit", "dft"):
            d.pop(k, None)
        hit = fp_lookup(d["n"])
        if not hit or hit[0] in used:
            continue
        s, v = hit
        used.add(s)
        matched += 1
        rv = relv.get(s, 0)
        d["fpvol"] = v["vol"]; d["fpunits"] = v["units"]; d["fpprice"] = round(v["avgp"], 1); d["fpvel"] = rv
        dec_pct = (1 - (d["dec"] - 1) / 9.0) if d.get("dec") else 0.4  # 1=best ->1.0; unknown ->0.4
        # JB premium fit (price-led) — qualifies as a target only above a premium floor
        if v["avgp"] >= 32 and v["vol"] >= 200000:
            d["jbfit"] = round(100 * (0.45 * pP(v["avgp"]) + 0.28 * pV(v["vol"]) + 0.15 * pVel(rv) + 0.12 * dec_pct))
        # Dragonfly value fit (units-led, value price)
        if v["units"] >= 5000:
            d["dflyfit"] = round(100 * (0.50 * pU(v["units"]) + 0.30 * (1 - pP(v["avgp"])) + 0.20 * pVel(rv)))

    # Tier each brand by quantile AMONG ITS QUALIFIERS so we get a real T1/T2/T3 spread.
    def assign_tiers(score_key, tier_key):
        qs = sorted((d[score_key] for d in data if d.get(score_key) is not None), reverse=True)
        if not qs:
            return {1: 0, 2: 0, 3: 0}
        c1 = qs[int(len(qs) * 0.20)]          # top 20% -> T1
        c2 = qs[min(len(qs) - 1, int(len(qs) * 0.55))]  # next 35% -> T2, rest -> T3
        cnt = {1: 0, 2: 0, 3: 0}
        for d in data:
            sc = d.get(score_key)
            if sc is None:
                continue
            t = 1 if sc >= c1 else (2 if sc >= c2 else 3)
            d[tier_key] = t; cnt[t] += 1
        return cnt

    jbc = assign_tiers("jbfit", "jbt")
    dfc = assign_tiers("dflyfit", "dft")
    print(f"\nmatched {matched}/{len(data)} doors to F+PR data | market median velocity {medv}%")
    print(f"JB premium tiers:  T1={jbc[1]}  T2={jbc[2]}  T3={jbc[3]}")
    print(f"Dfly value tiers:  T1={dfc[1]}  T2={dfc[2]}  T3={dfc[3]}")
    ranked = sorted([d for d in data if d.get("jbt")], key=lambda d: -d["jbfit"])
    print("\nTop 15 JB premium targets (data-driven tier):")
    for d in ranked[:15]:
        cust = " [already JB]" if "JB Tier" in d.get("role", "") else ""
        print(f"  T{d['jbt']} fit{d['jbfit']:>3}  {d['n'][:30]:<30} ${d['fpvol']/1000:>5.0f}k F+PR  ${d['fpprice']:>4.0f} avg  vel{d['fpvel']:+d}%  [{d['role']}]{cust}")

    if "--dry" in sys.argv:
        print("\n--dry: no write."); return
    new = html[:m.start(1)] + json.dumps(data, ensure_ascii=False) + html[m.end(1):]
    open(HTML, "w", encoding="utf-8").write(new)
    print(f"\nWrote F+PR fit (jbt/dft) into index.html. Next: python tools/sync_accounts.py")


if __name__ == "__main__":
    main()
