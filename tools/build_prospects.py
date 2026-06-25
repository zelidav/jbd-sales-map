#!/usr/bin/env python3
"""Surface top off-map performers as prospects on the map.

A "prospect" = a store that ranks in the Pistil sales export but is NOT one of our
mapped doors. Being in the rank export is proof it is operating, so we do NOT filter
by OCM operational_status here — we only need OCM (or a web-resolved override) for a
street address + license.

Pipeline:
  off-map rank rows -> resolve address (tools/prospect_addresses.json override first,
  else unique OCM dba/entity name match) -> geocode (Nominatim, cached) -> append to
  DATA as role "New Prospect" with prospect=true, the rank/momentum fields, and op.

Re-runnable. Geocodes are cached in tools/_cache/geocode.json.

Usage: python tools/build_prospects.py A.xlsx B.xlsx C.xlsx   (30/90/180-day, any order)
       python tools/build_prospects.py                         (newest 3 in ~/Downloads)
"""
import re, os, sys, json, glob, time, unicodedata, urllib.request, urllib.parse
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
HTML = os.path.join(ROOT, "index.html")
CACHE = os.path.join(ROOT, "tools", "_cache")
STOP = set("the llc inc co of a and an at to ny nyc rec dispensary dispensaries store shop adult use".split())
RETAIL = {"OCMRETL", "OCMCAURD22", "OCMXROD", "OCMMICR"}


def cset(s):
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().lower()
    s = re.sub(r"\(.*?\)", " ", s)
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return frozenset(t for t in s.split() if t and t not in STOP)


def cbase(s):
    return cset(str(s or "").split(" - ")[0])


def load_rank(p):
    wb = openpyxl.load_workbook(p, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        out[r[1]] = {"rank": int(r[0]), "store": r[1], "units": int(r[2] or 0), "vol": int(r[4] or 0)}
    return out


def geocode(query, cache):
    if query in cache:
        return cache[query]
    url = "https://nominatim.openstreetmap.org/search?" + urllib.parse.urlencode(
        {"format": "json", "limit": 1, "countrycodes": "us", "q": query})
    req = urllib.request.Request(url, headers={"User-Agent": "jbd-sales-map/1.0 (prospect geocoder)"})
    try:
        with urllib.request.urlopen(req, timeout=20) as r:
            j = json.load(r)
        cache[query] = [float(j[0]["lat"]), float(j[0]["lon"])] if j else None
    except Exception as e:
        print(f"  geocode error for {query!r}: {e}")
        cache[query] = None
    time.sleep(1.1)  # Nominatim politeness
    return cache[query]


def main():
    pos = [a for a in sys.argv[1:] if not a.startswith("--")]
    files = pos[:3] if len(pos) >= 3 else sorted(
        glob.glob(os.path.join(os.path.expanduser("~"), "Downloads", "store_rank_*.xlsx")),
        key=os.path.getmtime)[-3:]
    wins = sorted(((f, load_rank(f)) for f in files), key=lambda fr: sum(x["vol"] for x in fr[1].values()))
    (_, d30), (_, d90), (_, d180) = wins
    med = json.load(open(os.path.join(CACHE, "market_baseline.json"))).get("mom_median", 34)

    def metrics(store):
        r90 = d90.get(store)
        if not r90:
            return None
        m = {"psr": r90["rank"], "svol": r90["vol"], "sunits": r90["units"]}
        r30 = d30.get(store); r180 = d180.get(store)
        if r30:
            m["svol30"] = r30["vol"]
        if r180:
            m["svol180"] = r180["vol"]
        if r30 and r90["vol"]:
            m["mom"] = round(100 * (r30["vol"] / (r90["vol"] / 3.0) - 1))
            m["momr"] = m["mom"] - med
        if r90["vol"] and r180:
            m["trend"] = round(100 * ((r90["vol"] / 3.0) / (r180["vol"] / 6.0) - 1))
        return m

    ocm = json.load(open(os.path.join(CACHE, "ocm_full.json")))
    overrides = {k: v for k, v in json.load(open(os.path.join(ROOT, "tools", "prospect_addresses.json"))).items()
                 if not k.startswith("_")}
    html = open(HTML, encoding="utf-8").read()
    m = re.search(r"var DATA=(\[.*?\]);", html, re.S)
    data = json.loads(m.group(1))
    # strip any prior prospects FIRST so re-runs are idempotent (they must not count
    # as "already on map" or we'd shrink the set each run).
    data = [d for d in data if not d.get("prospect")]
    map_lic = {d.get("lic") for d in data if d.get("lic")}

    # door name-sets to detect "already on map"
    mapsets = set()
    for d in data:
        x = ocm.get(d.get("lic")) or {}
        for nm in (d.get("n"), x.get("dba"), x.get("ent")):
            mapsets.add(cset(nm)); mapsets.add(cbase(nm))

    def on_map(store):
        return cset(store) in mapsets or cbase(store) in mapsets

    # OCM retail name index (unique sets only -> avoid ambiguous collisions)
    ocm_sets = {}
    for x in ocm.values():
        if x.get("type") not in RETAIL or not x.get("a"):
            continue
        for nm in (x.get("dba"), x.get("ent")):
            cs = cset(nm)
            if cs:
                ocm_sets.setdefault(cs, x)

    def ocm_match(store):
        for key in (cset(store), cbase(store)):
            if key in ocm_sets:
                return ocm_sets[key]
        return None

    geocache = {}
    gpath = os.path.join(CACHE, "geocode.json")
    if os.path.exists(gpath):
        geocache = json.load(open(gpath))

    offmap = [r for r in sorted(d90.values(), key=lambda x: x["rank"]) if not on_map(r["store"])]
    prospects = []
    seen_lic = set(map_lic)
    seen_addr = set()
    skipped = []
    for r in offmap:
        store = r["store"]
        ov = overrides.get(store)
        if ov:
            addr = {k: ov.get(k, "") for k in ("a", "c", "co", "rg", "zip")}
            lic = ov.get("lic", "")
            op = (ocm.get(lic) or {}).get("op", "Active") if lic else "Active"
            web = ""
        else:
            x = ocm_match(store)
            if not x:
                skipped.append(r); continue
            lic = x["lic"]
            addr = {"a": x.get("a", ""), "c": x.get("c", ""), "co": x.get("co", ""),
                    "rg": x.get("rg", ""), "zip": x.get("zip", "")}
            op = x.get("op", "Active")
            web = x.get("web", "")
        if not addr["a"]:
            skipped.append(r); continue
        akey = (addr["a"].lower(), addr["c"].lower())
        if (lic and lic in seen_lic) or akey in seen_addr:
            continue
        if lic:
            seen_lic.add(lic)
        seen_addr.add(akey)
        q = ", ".join([addr["a"], addr["c"], "NY " + addr["zip"]]).strip()
        ll = geocode(q, geocache)
        if not ll:
            skipped.append(r); continue
        rec = {"n": store.replace(" (Rec)", "").strip(), "lat": ll[0], "lng": ll[1],
               "a": addr["a"], "c": addr["c"], "co": addr["co"], "rg": addr["rg"],
               "nb": "", "role": "New Prospect", "ds": "", "days": None, "rev": None,
               "dec": None, "tier": "", "rep": "", "poc": "", "ph": "", "lic": lic,
               "op": op, "opened": "", "prospect": True, "src": "rank"}
        rec.update(metrics(store) or {})
        if web:
            rec["web"] = web
        prospects.append(rec)

    json.dump(geocache, open(gpath, "w"))
    print(f"off-map rank rows: {len(offmap)} | resolved+geocoded prospects: {len(prospects)} | unresolved: {len(skipped)}")
    print("\nAdded prospects (top 25 by rank):")
    for p in sorted(prospects, key=lambda x: x.get("psr", 9999))[:25]:
        print(f"  #{p.get('psr','?'):>3} ${p.get('svol',0):>9,} {p['n'][:30]:<30} {p['a'][:22]:<22} {p['c']}  momr={p.get('momr','')}")

    if "--dry" in sys.argv:
        print("\n--dry: no write.")
        return
    data += prospects  # prior prospects already stripped above
    new = html[:m.start(1)] + json.dumps(data, ensure_ascii=False) + html[m.end(1):]
    open(HTML, "w", encoding="utf-8").write(new)
    print(f"\nWrote {len(prospects)} prospects into index.html (total doors now {len(data)}).")
    print("Next: python tools/sync_accounts.py")


if __name__ == "__main__":
    main()
