#!/usr/bin/env python3
"""Attach Pistil Store Rank + momentum/trend to the map's DATA from THREE windows.

Pistil store-ranking exports key on trade/DBA store name (no license #). We bridge
to map doors via the OCM dataset (dba + entity_name per license):
rank.store -> OCM dba / entity / map name -> license -> map door. Match is
high-confidence only (exact normalized token-set, word-order independent, or a
strict subset) — NO loose fuzzy.

Three trailing windows (auto-detected by total volume: smallest=30d, mid=90d,
largest=180d):
  - 180-day  = headline "Pistil Store Rank" (most stable).  -> psr, svol, sunits
  - 90-day, 30-day                                          -> svol90, svol30
Derived sales intelligence (monthly run-rate = window sales / months):
  - mom   = recent momentum  = round(100*(svol30/(svol90/3) - 1))   # 30d vs 90d pace
  - trend = medium trend     = round(100*(svol90/(svol180/2) - 1))  # 90d vs 180d pace
  +N% = running hotter than the longer window (accelerating); -N% = cooling.

Usage:
  python tools/build_store_rank.py A.xlsx B.xlsx C.xlsx   # order-independent
  python tools/build_store_rank.py                         # newest 3 store_rank_*.xlsx in ~/Downloads
"""
import re, os, sys, json, glob, unicodedata
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
HTML = os.path.join(ROOT, "index.html")
CACHE = os.path.join(ROOT, "tools", "_cache", "ocm_full.json")
STOP = set("the llc inc co of a and an at to ny nyc rec dispensary dispensaries store shop adult use".split())


def cset(s):
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().lower()
    s = re.sub(r"\(.*?\)", " ", s)
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    return frozenset(t for t in s.split() if t and t not in STOP)


def cbase(s):
    return cset(str(s or "").split(" - ")[0])


def window_days(avgs):
    """Recover a Pistil export's true window length in days.

    Pistil exports carry NO date metadata, so the window was previously *guessed*
    from total volume — which silently mislabelled a 69-day file as "90-day" and a
    93-day file as "180-day", corrupting every momentum figure on the map.
    'Avg # SKUs Stocked' is a per-day mean, so its denominator IS the day count:
    find the smallest d that makes avg*d an integer for the overwhelming majority
    of rows. Verified exact against every historical export.
    """
    frac = [v for v in avgs if isinstance(v, (int, float)) and v != int(v)]
    if not frac:
        return None
    for d in range(1, 200):
        ok = sum(1 for v in frac if abs(v * d - round(v * d)) < 1e-3)
        if ok / len(frac) > 0.85:
            return d
    return None


def load_rank(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out, avgs = {}, []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        out[r[1]] = {"rank": int(r[0]), "store": r[1], "units": int(r[2] or 0), "vol": int(r[4] or 0)}
        if len(avgs) < 120 and len(r) > 5:
            avgs.append(r[5])
    days = window_days(avgs)
    if not days:
        raise SystemExit(f"Cannot determine window length for {os.path.basename(path)} — refusing to guess.")
    for v in out.values():
        v["days"] = days
    return out, days


def main():
    pos = [a for a in sys.argv[1:] if not a.startswith("--")]
    files = pos[:3] if len(pos) >= 3 else sorted(
        glob.glob(os.path.join(os.path.expanduser("~"), "Downloads", "store_rank_*.xlsx")),
        key=os.path.getmtime)[-3:]
    if len(files) < 3:
        raise SystemExit("Need three store_rank_*.xlsx files (30/90/180-day).")
    loaded = []
    for f in files:
        rows, days = load_rank(f)
        loaded.append((f, rows, days))
    loaded.sort(key=lambda fr: fr[2])  # ascending by MEASURED window length
    (fS, dS, nS), (fM, dM, nM), (fL, dL, nL) = loaded
    if len({nS, nM, nL}) < 3:
        raise SystemExit(f"Windows are not distinct ({nS}/{nM}/{nL} days) — pull three different ranges.")
    print("windows (measured from the files, not guessed):")
    for f, rows, days in loaded:
        tot = sum(x["vol"] for x in rows.values())
        print(f"  {days:>4}d  {os.path.basename(f):42} rows={len(rows):4} "
              f"total=${tot:,} (${tot/days:,.0f}/day)")

    html = open(HTML, encoding="utf-8").read()
    m = re.search(r"var DATA=(\[.*?\]);", html, re.S)
    data = json.loads(m.group(1))
    ocm = json.load(open(CACHE)) if os.path.exists(CACHE) else {}

    door_set, door_join = {}, {}
    for d in data:
        x = ocm.get(d.get("lic")) or {}
        for nm in (d.get("n"), x.get("dba"), x.get("ent")):
            for v in (cset(nm), cbase(nm)):
                if v:
                    door_set.setdefault(v, d)
                    door_join.setdefault("".join(sorted(v)), d)

    def find_door(store):
        for key in (cset(store), cbase(store)):
            if key in door_set:
                return door_set[key]
        for key in (cset(store), cbase(store)):
            j = "".join(sorted(key))
            if j and j in door_join:
                return door_join[j]
        return None

    for d in data:
        for k in ("psr", "svol", "sunits", "svol30", "svol180", "mom", "trend", "momr"):
            d.pop(k, None)

    def attach(window, vol_key, rank_key=None, units_key=None):
        seen = set()
        for r in sorted(window.values(), key=lambda x: x["rank"]):
            d = find_door(r["store"])
            if not d:
                continue
            lic = d.get("lic") or d.get("n")
            if lic in seen:
                continue
            seen.add(lic)
            d[vol_key] = r["vol"]
            if rank_key:
                d[rank_key] = r["rank"]
            if units_key:
                d[units_key] = r["units"]
        return len(seen)

    matched = attach(dM, "svol", "psr", "sunits")  # headline = mid window
    attach(dS, "svol30")    # legacy key = SHORT window volume
    attach(dL, "svol180")   # legacy key = LONG window volume

    def pct(recent_rate, prior_rate):
        """Percent change of a daily run-rate vs the daily run-rate that preceded it."""
        if not prior_rate or prior_rate <= 0 or recent_rate is None:
            return None
        return round(100 * (recent_rate / prior_rate - 1))

    momn = 0
    for d in data:
        vS, vM, vL = d.get("svol30"), d.get("svol"), d.get("svol180")
        # Compare the recent window against the period IMMEDIATELY BEFORE it, not
        # against a longer window that CONTAINS it. Windows are nested (the short
        # window's sales are also inside the mid window), so the prior period is the
        # non-overlapping remainder: (vol_mid - vol_short) over (days_mid - days_short).
        if vS and vM and vM > vS:
            d["mom"] = pct(vS / nS, (vM - vS) / (nM - nS))
        if vM and vL and vL > vM:
            d["trend"] = pct(vM / nM, (vL - vM) / (nL - nM))
        if d.get("mom") is not None:
            momn += 1
    # The whole NY market is growing, so raw momentum is positive almost everywhere.
    # Market-relativize: momr = store momentum minus the STATEWIDE median (computed over
    # every store in the rank export, not just mapped ones), so a rep sees who is
    # accelerating FASTER (or slower) than the typical NY store. Persisted to _cache so
    # build_prospects.py uses the identical baseline.
    mkt = []
    for st, rM in dM.items():
        rS = dS.get(st)
        if rS and rM["vol"] > rS["vol"]:
            v = pct(rS["vol"] / nS, (rM["vol"] - rS["vol"]) / (nM - nS))
            if v is not None:
                mkt.append(v)
    mkt.sort()
    med = mkt[len(mkt) // 2] if mkt else 0
    json.dump({"mom_median": med, "win_short": nS, "win_mid": nM, "win_long": nL},
              open(os.path.join(ROOT, "tools", "_cache", "market_baseline.json"), "w"))
    for d in data:
        if d.get("mom") is not None:
            d["momr"] = d["mom"] - med
    print(f"statewide median momentum (last {nS}d vs prior {nM-nS}d): {med}%  — momr is relative to this")
    if abs(med) > 25:
        print(f"  ! median momentum of {med}% is large; check the windows are what you expect.")

    ranked = sorted([d for d in data if d.get("psr")], key=lambda d: d["psr"])
    print(f"\nmatched {matched}/{len(data)} doors to a {nM}-day rank; momentum on {momn}.")
    print(f"Top 12 ({nM}d rank · last {nS}d vs prior {nM-nS}d · last {nM}d vs prior {nL-nM}d):")
    for d in ranked[:12]:
        mm = d.get("mom"); tr = d.get("trend")
        f = lambda v: "  n/a" if v is None else f"{'+' if v >= 0 else ''}{v}%"
        print(f"  #{d['psr']:>3}  {d['n'][:30]:<30} ${d.get('svol',0):>10,}  mom {f(mm):>5}  trend {f(tr):>5}")

    if "--dry" in sys.argv:
        print("\n--dry: no write.")
        return
    new = html[:m.start(1)] + json.dumps(data, ensure_ascii=False) + html[m.end(1):]
    # Publish the measured window lengths so the UI labels periods honestly instead of
    # hard-coding "30/90/180d" (svol30/svol/svol180 are short/mid/long, not fixed spans).
    wj = json.dumps({"s": nS, "m": nM, "l": nL})
    if re.search(r"var WINDOWS=\{.*?\};", new):
        new = re.sub(r"var WINDOWS=\{.*?\};", f"var WINDOWS={wj};", new, count=1)
    else:
        new = new.replace("var DATA=", f"var WINDOWS={wj};\nvar DATA=", 1)
    open(HTML, "w", encoding="utf-8").write(new)
    print(f"\nWrote ranks+momentum into index.html. Next: python tools/sync_accounts.py")


if __name__ == "__main__":
    main()
